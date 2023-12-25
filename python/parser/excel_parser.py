from openpyxl import load_workbook
import sys
from io import BytesIO


class HuExcelParser:
    def __call__(self, fnm):
        if isinstance(fnm, str):wb = load_workbook(fnm)
        else: wb = load_workbook(BytesIO(fnm))
        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            lines = []
            for r in ws.rows:
                lines.append(
                    "\t".join([str(c.value) if c.value is not None else "" for c in r]))
            res.append(f"《{sheetname}》\n" + "\n".join(lines))
        return res


if __name__ == "__main__":
    psr = HuExcelParser()
    psr(sys.argv[1])
