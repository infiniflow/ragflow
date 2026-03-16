#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

import logging
import re
import sys
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from PIL import Image

from rag.nlp import find_codec

# copied from `/openpyxl/cell/cell.py`
ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


class RAGFlowExcelParser:
    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        # Read first 4 bytes to determine file type
        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        if not (file_head.startswith(b"PK\x03\x04") or file_head.startswith(b"\xd0\xcf\x11\xe0")):
            logging.info("Not an Excel file, converting CSV to Excel Workbook")

            try:
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object, on_bad_lines='skip')
                return RAGFlowExcelParser._dataframe_to_workbook(df)

            except Exception as e_csv:
                raise Exception(f"Failed to parse CSV and convert to Excel Workbook: {e_csv}")

        try:
            return load_workbook(file_like_object, data_only=True)
        except Exception as e:
            logging.info(f"openpyxl load error: {e}, try pandas instead")
            try:
                file_like_object.seek(0)
                try:
                    dfs = pd.read_excel(file_like_object, sheet_name=None)
                    return RAGFlowExcelParser._dataframe_to_workbook(dfs)
                except Exception as ex:
                    logging.info(f"pandas with default engine load error: {ex}, try calamine instead")
                    file_like_object.seek(0)
                    df = pd.read_excel(file_like_object, engine="calamine")
                    return RAGFlowExcelParser._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"pandas.read_excel error: {e_pandas}, original openpyxl error: {e}")

    @staticmethod
    def _clean_dataframe(df: pd.DataFrame):
        def clean_string(s):
            if isinstance(s, str):
                return ILLEGAL_CHARACTERS_RE.sub(" ", s)
            return s

        return df.apply(lambda col: col.map(clean_string))

    @staticmethod
    def _dataframe_to_workbook(df):
        # if contains multiple sheets use _dataframes_to_workbook
        if isinstance(df, dict) and len(df) > 1:
            return RAGFlowExcelParser._dataframes_to_workbook(df)

        df = RAGFlowExcelParser._clean_dataframe(df)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        return wb
    
    @staticmethod
    def _dataframes_to_workbook(dfs: dict):
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        for sheet_name, df in dfs.items():
            df = RAGFlowExcelParser._clean_dataframe(df)
            ws = wb.create_sheet(title=sheet_name)
            for col_num, column_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_name)
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        return wb

    @staticmethod
    def _extract_images_from_worksheet(ws, sheetname=None):
        """
        Extract images from a worksheet and enrich them with vision-based descriptions.

        Returns: List[dict]
        """
        images = getattr(ws, "_images", [])
        if not images:
            return []

        raw_items = []

        for img in images:
            try:
                img_bytes = img._data()
                pil_img = Image.open(BytesIO(img_bytes)).convert("RGB")

                anchor = img.anchor
                if hasattr(anchor, "_from") and hasattr(anchor, "_to"):
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = anchor._to.row + 1, anchor._to.col + 1
                    if r1 == r2 and c1 == c2:
                        span = "single_cell"
                    else:
                        span = "multi_cell"
                else:
                    r1, c1 = anchor._from.row + 1, anchor._from.col + 1
                    r2, c2 = r1, c1
                    span = "single_cell"

                item = {
                    "sheet": sheetname or ws.title,
                    "image": pil_img,
                    "image_description": "",
                    "row_from": r1,
                    "col_from": c1,
                    "row_to": r2,
                    "col_to": c2,
                    "span_type": span,
                }
                raw_items.append(item)
            except Exception:
                continue
        return raw_items

    @staticmethod
    def _get_actual_row_count(ws):
        max_row = ws.max_row
        if not max_row:
            return 0
        if max_row <= 10000:
            return max_row

        max_col = min(ws.max_column or 1, 50)

        def row_has_data(row_idx):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    return True
            return False

        if not any(row_has_data(i) for i in range(1, min(101, max_row + 1))):
            return 0

        left, right = 1, max_row
        last_data_row = 1

        while left <= right:
            mid = (left + right) // 2
            found = False
            for r in range(mid, min(mid + 10, max_row + 1)):
                if row_has_data(r):
                    found = True
                    last_data_row = max(last_data_row, r)
                    break
            if found:
                left = mid + 1
            else:
                right = mid - 1

        for r in range(last_data_row, min(last_data_row + 500, max_row + 1)):
            if row_has_data(r):
                last_data_row = r

        return last_data_row

    @staticmethod
    def _get_rows_limited(ws):
        actual_rows = RAGFlowExcelParser._get_actual_row_count(ws)
        if actual_rows == 0:
            return []
        return list(ws.iter_rows(min_row=1, max_row=actual_rows))

    def html(self, fnm, chunk_rows=256):
        from html import escape

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        tb_chunks = []

        def _fmt(v):
            if v is None:
                return ""
            return str(v).strip()

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                rows = RAGFlowExcelParser._get_rows_limited(ws)
            except Exception as e:
                logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                continue

            if not rows:
                continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{escape(_fmt(t.value))}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(rows[1 + chunk_i * chunk_rows : min(1 + (chunk_i + 1) * chunk_rows, len(rows))]):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{escape(_fmt(c.value))}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def markdown(self, fnm):
        import pandas as pd

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        try:
            file_like_object.seek(0)
            df = pd.read_excel(file_like_object)
        except Exception as e:
            logging.warning(f"Parse spreadsheet error: {e}, trying to interpret as CSV file")
            file_like_object.seek(0)
            df = pd.read_csv(file_like_object, on_bad_lines='skip')
        df = df.replace(r"^\s*$", "", regex=True)
        return df.to_markdown(index=False)

    def __call__(self, fnm):
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)

        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            try:
                rows = RAGFlowExcelParser._get_rows_limited(ws)
            except Exception as e:
                logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                continue
            if not rows:
                continue
            ti = list(rows[0])
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                if not fields:
                    continue
                line = "; ".join(fields)
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0

            for sheetname in wb.sheetnames:
                try:
                    ws = wb[sheetname]
                    total += RAGFlowExcelParser._get_actual_row_count(ws)
                except Exception as e:
                    logging.warning(f"Skip sheet '{sheetname}' due to rows access error: {e}")
                    continue
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
