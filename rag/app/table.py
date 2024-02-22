#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
import copy
import re
from io import BytesIO
from xpinyin import Pinyin
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from dateutil.parser import parse as datetime_parse

from api.db.services.knowledgebase_service import KnowledgebaseService
from deepdoc.parser import is_english, tokenize
from rag.nlp import huqie
from deepdoc.parser import ExcelParser


class Excel(ExcelParser):
    def __call__(self, fnm, binary=None, callback=None):
        if not binary:
            wb = load_workbook(fnm)
        else:
            wb = load_workbook(BytesIO(binary))
        total = 0
        for sheetname in wb.sheetnames:
            total += len(list(wb[sheetname].rows))

        res, fails, done = [], [], 0
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            headers = [cell.value for cell in rows[0]]
            missed = set([i for i, h in enumerate(headers) if h is None])
            headers = [
                cell.value for i,
                cell in enumerate(
                    rows[0]) if i not in missed]
            data = []
            for i, r in enumerate(rows[1:]):
                row = [
                    cell.value for ii,
                    cell in enumerate(r) if ii not in missed]
                if len(row) != len(headers):
                    fails.append(str(i))
                    continue
                data.append(row)
                done += 1
                if done % 999 == 0:
                    callback(done * 0.6 / total, ("Extract records: {}".format(len(res)) + (
                        f"{len(fails)} failure({sheetname}), line: %s..." % (",".join(fails[:3])) if fails else "")))
            res.append(pd.DataFrame(np.array(data), columns=headers))

        callback(0.6, ("Extract records: {}. ".format(done) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))
        return res


def trans_datatime(s):
    try:
        return datetime_parse(s.strip()).strftime("%Y-%m-%dT%H:%M:%S")
    except Exception as e:
        pass


def trans_bool(s):
    if re.match(r"(true|yes|是)$", str(s).strip(), flags=re.IGNORECASE):
        return ["yes", "是"]
    if re.match(r"(false|no|否)$", str(s).strip(), flags=re.IGNORECASE):
        return ["no", "否"]


def column_data_type(arr):
    uni = len(set([a for a in arr if a is not None]))
    counts = {"int": 0, "float": 0, "text": 0, "datetime": 0, "bool": 0}
    trans = {t: f for f, t in
             [(int, "int"), (float, "float"), (trans_datatime, "datetime"), (trans_bool, "bool"), (str, "text")]}
    for a in arr:
        if a is None:
            continue
        if re.match(r"[+-]?[0-9]+(\.0+)?$", str(a).replace("%%", "")):
            counts["int"] += 1
        elif re.match(r"[+-]?[0-9.]+$", str(a).replace("%%", "")):
            counts["float"] += 1
        elif re.match(r"(true|false|yes|no|是|否)$", str(a), flags=re.IGNORECASE):
            counts["bool"] += 1
        elif trans_datatime(str(a)):
            counts["datetime"] += 1
        else:
            counts["text"] += 1
    counts = sorted(counts.items(), key=lambda x: x[1] * -1)
    ty = counts[0][0]
    for i in range(len(arr)):
        if arr[i] is None:
            continue
        try:
            arr[i] = trans[ty](str(arr[i]))
        except Exception as e:
            arr[i] = None
    if ty == "text":
        if len(arr) > 128 and uni / len(arr) < 0.1:
            ty = "keyword"
    return arr, ty


def chunk(filename, binary=None, callback=None, **kwargs):
    """
        Excel and csv(txt) format files are supported.
        For csv or txt file, the delimiter between columns is TAB.
        The first line must be column headers.
        Column headers must be meaningful terms inorder to make our NLP model understanding.
        It's good to enumerate some synonyms using slash '/' to separate, and even better to
        enumerate values using brackets like 'gender/sex(male, female)'.
        Here are some examples for headers:
            1. supplier/vendor\tcolor(yellow, red, brown)\tgender/sex(male, female)\tsize(M,L,XL,XXL)
            2. 姓名/名字\t电话/手机/微信\t最高学历（高中，职高，硕士，本科，博士，初中，中技，中专，专科，专升本，MPA，MBA，EMBA）

        Every row in table will be treated as a chunk.
    """

    if re.search(r"\.xlsx?$", filename, re.IGNORECASE):
        callback(0.1, "Start to parse.")
        excel_parser = Excel()
        dfs = excel_parser(filename, binary, callback)
    elif re.search(r"\.(txt|csv)$", filename, re.IGNORECASE):
        callback(0.1, "Start to parse.")
        txt = ""
        if binary:
            txt = binary.decode("utf-8")
        else:
            with open(filename, "r") as f:
                while True:
                    l = f.readline()
                    if not l:
                        break
                    txt += l
        lines = txt.split("\n")
        fails = []
        headers = lines[0].split(kwargs.get("delimiter", "\t"))
        rows = []
        for i, line in enumerate(lines[1:]):
            row = [l for l in line.split(kwargs.get("delimiter", "\t"))]
            if len(row) != len(headers):
                fails.append(str(i))
                continue
            rows.append(row)
            if len(rows) % 999 == 0:
                callback(len(rows) * 0.6 / len(lines), ("Extract records: {}".format(len(rows)) + (
                    f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))

        callback(0.6, ("Extract records: {}".format(len(rows)) + (
            f"{len(fails)} failure, line: %s..." % (",".join(fails[:3])) if fails else "")))

        dfs = [pd.DataFrame(np.array(rows), columns=headers)]

    else:
        raise NotImplementedError(
            "file type not supported yet(excel, text, csv supported)")

    res = []
    PY = Pinyin()
    fieds_map = {
        "text": "_tks",
        "int": "_int",
        "keyword": "_kwd",
        "float": "_flt",
        "datetime": "_dt",
        "bool": "_kwd"}
    for df in dfs:
        for n in ["id", "_id", "index", "idx"]:
            if n in df.columns:
                del df[n]
        clmns = df.columns.values
        txts = list(copy.deepcopy(clmns))
        py_clmns = [PY.get_pinyins(re.sub(r"(/.*|（[^（）]+?）|\([^()]+?\))", "", n), '_')[0] for n in clmns]
        clmn_tys = []
        for j in range(len(clmns)):
            cln, ty = column_data_type(df[clmns[j]])
            clmn_tys.append(ty)
            df[clmns[j]] = cln
            if ty == "text":
                txts.extend([str(c) for c in cln if c])
        clmns_map = [(py_clmns[j] + fieds_map[clmn_tys[j]], clmns[j])
                     for i in range(len(clmns))]

        eng = is_english(txts)
        for ii, row in df.iterrows():
            d = {}
            row_txt = []
            for j in range(len(clmns)):
                if row[clmns[j]] is None:
                    continue
                fld = clmns_map[j][0]
                d[fld] = row[clmns[j]] if clmn_tys[j] != "text" else huqie.qie(
                    row[clmns[j]])
                row_txt.append("{}:{}".format(clmns[j], row[clmns[j]]))
            if not row_txt:
                continue
            tokenize(d, "; ".join(row_txt), eng)
            res.append(d)

        KnowledgebaseService.update_parser_config(
            kwargs["kb_id"], {"field_map": {k: v for k, v in clmns_map}})
    callback(0.6, "")

    return res


if __name__ == "__main__":
    import sys

    def dummy(a, b):
        pass

    chunk(sys.argv[1], callback=dummy)
