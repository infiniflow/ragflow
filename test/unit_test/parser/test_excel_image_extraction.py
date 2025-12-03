#
#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

"""
Unit tests for Excel image extraction functionality
"""

import pytest
from deepdoc.parser.excel_parser import RAGFlowExcelParser
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
from PIL import Image
import base64


class TestExcelImageExtraction:
    """Test Excel image extraction functionality"""
    
    @pytest.fixture
    def sample_excel_with_image(self):
        """Create a sample Excel file with an embedded image"""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Add some data
        ws['A1'] = "Header"
        ws['B1'] = "Data"
        ws['A2'] = "Row 1"
        ws['B2'] = 100
        
        # Create a simple test image (1x1 red pixel)
        img = Image.new('RGB', (10, 10), color='red')
        img_buffer = BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        # Add image to worksheet
        openpyxl_img = OpenpyxlImage(img_buffer)
        openpyxl_img.anchor = 'D2'  # Position at cell D2
        ws.add_image(openpyxl_img)
        
        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
    
    def test_extract_images_from_excel(self, sample_excel_with_image):
        """Test extracting images from Excel file"""
        parser = RAGFlowExcelParser()
        
        images = parser.extract_images(sample_excel_with_image)
        
        # Should have extracted one image
        assert len(images) == 1
        
        # Check image properties
        img = images[0]
        assert 'image_data' in img
        assert 'format' in img
        assert 'sheet' in img
        assert 'anchor' in img
        assert 'description' in img
        assert 'size' in img
        assert 'index' in img
        
        # Verify sheet name
        assert img['sheet'] == 'TestSheet'
        
        # Verify format
        assert img['format'] in ['png', 'jpeg', 'jpg', 'gif', 'bmp']
        
        # Verify image data is base64 encoded
        assert isinstance(img['image_data'], str)
        try:
            base64.b64decode(img['image_data'])
        except Exception:
            pytest.fail("Image data is not valid base64")
    
    def test_extract_images_from_excel_without_images(self):
        """Test extracting images from Excel file without images"""
        parser = RAGFlowExcelParser()
        
        # Create simple Excel without images
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test"
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        images = parser.extract_images(excel_buffer.getvalue())
        
        # Should have no images
        assert len(images) == 0
    
    def test_extract_images_multiple_sheets(self):
        """Test extracting images from multiple sheets"""
        # Create workbook with multiple sheets
        wb = Workbook()
        
        # First sheet with image
        ws1 = wb.active
        ws1.title = "Sheet1"
        img1 = Image.new('RGB', (5, 5), color='blue')
        img_buffer1 = BytesIO()
        img1.save(img_buffer1, format='PNG')
        img_buffer1.seek(0)
        openpyxl_img1 = OpenpyxlImage(img_buffer1)
        ws1.add_image(openpyxl_img1, 'A1')
        
        # Second sheet with image
        ws2 = wb.create_sheet("Sheet2")
        img2 = Image.new('RGB', (5, 5), color='green')
        img_buffer2 = BytesIO()
        img2.save(img_buffer2, format='PNG')
        img_buffer2.seek(0)
        openpyxl_img2 = OpenpyxlImage(img_buffer2)
        ws2.add_image(openpyxl_img2, 'B2')
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        parser = RAGFlowExcelParser()
        images = parser.extract_images(excel_buffer.getvalue())
        
        # Should have extracted two images
        assert len(images) == 2
        
        # Verify different sheets
        sheet_names = {img['sheet'] for img in images}
        assert 'Sheet1' in sheet_names
        assert 'Sheet2' in sheet_names
    
    def test_column_letter_conversion(self):
        """Test column number to letter conversion"""
        assert RAGFlowExcelParser._number_to_column_letter(0) == 'A'
        assert RAGFlowExcelParser._number_to_column_letter(1) == 'B'
        assert RAGFlowExcelParser._number_to_column_letter(25) == 'Z'
        assert RAGFlowExcelParser._number_to_column_letter(26) == 'AA'
        assert RAGFlowExcelParser._number_to_column_letter(27) == 'AB'
    
    def test_extract_images_with_description(self, sample_excel_with_image):
        """Test that image descriptions are extracted"""
        parser = RAGFlowExcelParser()
        images = parser.extract_images(sample_excel_with_image)
        
        assert len(images) > 0
        # Description should not be empty
        assert images[0]['description']
        assert isinstance(images[0]['description'], str)
    
    def test_extract_images_with_size(self, sample_excel_with_image):
        """Test that image sizes are extracted"""
        parser = RAGFlowExcelParser()
        images = parser.extract_images(sample_excel_with_image)
        
        assert len(images) > 0
        # Size should be a tuple
        assert isinstance(images[0]['size'], tuple)
        assert len(images[0]['size']) == 2
        width, height = images[0]['size']
        assert width >= 0
        assert height >= 0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
