#
#  Copyright 2025 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

import importlib.util
import os
import sys
from io import BytesIO
from unittest import mock

import pytest

# Import RAGFlowExcelParser directly by file path to avoid triggering
# deepdoc/parser/__init__.py and rag.nlp, which pull in heavy dependencies.
for _m in ["rag.nlp", "rag.utils", "rag.utils.lazy_image"]:
    if _m not in sys.modules:
        sys.modules[_m] = mock.MagicMock()


def _find_project_root(marker="pyproject.toml"):
    d = os.path.dirname(os.path.abspath(__file__))
    while d != os.path.dirname(d):
        if os.path.exists(os.path.join(d, marker)):
            return d
        d = os.path.dirname(d)
    return None


_PROJECT_ROOT = _find_project_root()
_spec = importlib.util.spec_from_file_location(
    "deepdoc.parser.excel_parser",
    os.path.join(_PROJECT_ROOT, "deepdoc", "parser", "excel_parser.py"),
)
_mod = importlib.util.module_from_spec(_spec)
sys.modules["deepdoc.parser.excel_parser"] = _mod
_spec.loader.exec_module(_mod)

RAGFlowExcelParser = _mod.RAGFlowExcelParser


def _test_find_codec(binary):
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            binary.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("test", binary, 0, 1, "undecodable")


_mod.find_codec = _test_find_codec
_mod.decode_text = lambda binary, document_type="text": (binary.decode(_test_find_codec(binary)), _test_find_codec(binary))


def _make_xlsx(n_data_rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["H1", "H2"])
    for i in range(n_data_rows):
        ws.append([f"a{i}", f"b{i}"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _chunk_has_no_data_cells(chunk):
    return "<td>" not in chunk and "<td></td>" not in chunk


@pytest.mark.p2
def test_exact_multiple_does_not_emit_header_only_chunk():
    # 12 data rows with chunk_rows=12 (the value rag/app/naive.py uses).
    chunks = RAGFlowExcelParser().html(_make_xlsx(12), chunk_rows=12)
    assert len(chunks) == 1
    assert all(not _chunk_has_no_data_cells(c[0]) for c in chunks)
    assert chunks[0][1][0] == 0 and chunks[0][1][1] == 2 and chunks[0][1][2] == 13


@pytest.mark.p2
def test_multiple_of_chunk_rows_splits_without_spurious_chunk():
    # 24 data rows with chunk_rows=12 -> exactly 2 data chunks, no trailing header-only chunk.
    chunks = RAGFlowExcelParser().html(_make_xlsx(24), chunk_rows=12)
    assert len(chunks) == 2
    assert all(not _chunk_has_no_data_cells(c[0]) for c in chunks)
    assert chunks[0][1][1] == 2 and chunks[0][1][2] == 13
    assert chunks[1][1][1] == 14 and chunks[1][1][2] == 25


@pytest.mark.p2
def test_html_col_max_uses_widest_row():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["H1"])
    ws.append(["a", "b", "c"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    chunks = RAGFlowExcelParser().html(buf.read(), chunk_rows=12)
    assert chunks[0][1][3] == 1 and chunks[0][1][4] == 3


@pytest.mark.p2
def test_non_multiple_unchanged():
    # 13 data rows with chunk_rows=12 -> 2 chunks (12 + 1).
    chunks = RAGFlowExcelParser().html(_make_xlsx(13), chunk_rows=12)
    assert len(chunks) == 2
    assert all(not _chunk_has_no_data_cells(c[0]) for c in chunks)


def _make_xlsx_with_values(header, row):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(header)
    ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.mark.p2
def test_call_keeps_zero_valued_cells():
    # __call__ produces the text used for indexing. A numeric 0 (and 0.0 / False)
    # is real data, not an empty cell, so it must survive. The header is only
    # emitted alongside a kept value, so a dropped 0 also loses its "stock" label.
    lines = RAGFlowExcelParser()(_make_xlsx_with_values(["name", "stock"], ["widget", 0]))
    joined = " ".join(text for text, _ in lines)
    assert "stock" in joined and "0" in joined, lines
    assert lines[0][1][1] == 2


@pytest.mark.p2
def test_call_skips_truly_empty_cells():
    # None / empty-string cells carry no value and should still be skipped.
    lines = RAGFlowExcelParser()(_make_xlsx_with_values(["name", "note"], ["widget", None]))
    joined = " ".join(text for text, _ in lines)
    assert "note" not in joined, lines


@pytest.mark.p2
@pytest.mark.parametrize("encoding", ["utf-8", "utf-8-sig", "gbk", "gb18030"])
def test_csv_encoding_is_detected_without_losing_unicode(encoding):
    binary = "项目名称,备注\n核心系统重构,包含GBK语料导入测试\n".encode(encoding)
    lines = RAGFlowExcelParser()(binary)
    joined = " ".join(text for text, _ in lines)
    assert "项目名称" in joined
    assert "包含GBK语料导入测试" in joined


@pytest.mark.p2
def test_csv_accepts_gb18030_characters_not_supported_by_gbk():
    binary = "项目名称,备注\n扩展字符,𠀀\n".encode("gb18030")
    lines = RAGFlowExcelParser()(binary)
    assert "备注：𠀀" in lines[0][0]


@pytest.mark.p2
def test_corrupt_csv_encoding_still_fails():
    with pytest.raises(Exception, match="Failed to parse CSV"):
        RAGFlowExcelParser()(b"\xff\xfe\xfa\xfb")


def _make_two_sheet_xlsx():
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "s1"
    ws1.append(["H"])
    ws1.append(["a"])
    ws2 = wb.create_sheet("s2")
    ws2.append(["H"])
    ws2.append(["b"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.mark.p2
def test_call_emits_zero_based_sheet_index():
    # Flow JSON stores this index unchanged; add_positions later does pn+1 so
    # sheet 2 is stored as 2 and the preview selects xs.datas[1].
    lines = RAGFlowExcelParser()(_make_two_sheet_xlsx())
    sheets = {pos[0] for _, pos in lines}
    assert sheets == {0, 1}
    second = next(pos for _, pos in lines if pos[0] == 1)
    assert second == (1, 2, 2, 1, 1)
